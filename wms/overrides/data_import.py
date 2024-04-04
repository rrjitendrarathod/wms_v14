
import frappe
from frappe import _
from frappe.core.doctype.data_import.data_import import DataImport
# from frappe.core.doctype.data_import.importer import Importer
from frappe.utils.background_jobs import enqueue
from wms.overrides.custom_importer import CustomImporter
# from frappe.utils import formatdate
import datetime


class DataImportCustom(DataImport):    
    def start_import(self):
        from frappe.core.page.background_jobs.background_jobs import get_info
        from frappe.utils.scheduler import is_scheduler_inactive

        if is_scheduler_inactive() and not frappe.flags.in_test:
            frappe.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

        enqueued_jobs = [d.get("job_name") for d in get_info()]

        if self.name not in enqueued_jobs:
            enqueue(
                start_import,
                queue="default",
                timeout=10000,
                event="data_import",
                job_name=self.name,
                data_import=self.name,
                now=frappe.conf.developer_mode or frappe.flags.in_test,
            )
            return True

        return False
        
    # def export_errored_rows(self):
    #     return self.get_importer().export_errored_rows()
        
    def get_importer(self):
        return CustomImporter(self.reference_doctype, data_import=self)    

def start_import(data_import):
    """This method runs in background job"""
    data_import = frappe.get_doc("Data Import", data_import)
    
    try:
        i = CustomImporter(data_import.reference_doctype, data_import=data_import)     
        i.import_data()        
    except Exception:
        frappe.db.rollback()
        data_import.db_set("status", "Error")
        frappe.log_error(title=data_import.name)
        data_import.log_error("Data import failed")
    finally:
        frappe.flags.in_import = False

    return frappe.publish_realtime("data_import_refresh", {"data_import": data_import.name})  




@frappe.whitelist()
def update_import_file(data_import, import_file=None):
    if import_file:
        import csv
        from frappe.utils.file_manager import save_file
        
        file_doc = frappe.get_doc("File", {"file_url": import_file})
        file_path = file_doc.get_full_path()

        file_extension = file_doc.get_extension()[1]

        payor_conditions = {
                            "pps": "PPS - Non Medicare",
                            "non": "PPS - Non Medicare",
                            "apm": "Managed - Medicare - APM",
                            "mva": "Insurance - MVA",
                            "wco": "Insurance - WCO",

                            "managed-medicaid": "Managed - Medicaid",
                            "managedmedicaid": "Managed - Medicaid",
                            "managed medicaid": "Managed - Medicaid",
                            "managed–medicaid": "Managed - Medicaid",

                            "managed-medicare": "Managed - Medicare",
                            "managedmedicare": "Managed - Medicare",
                            "managed medicare": "Managed - Medicare",
                            "managed–medicare": "Managed - Medicare",

                            "contracts": "Contracts",
                            "commercial": "Commercial Insurance",
                            "insurance": "Commercial Insurance",
                            "self": "Self Pay - Home Health",
                            "pay": "Self Pay - Home Health",
                            "home": "Self Pay - Home Health",
                            "health": "Self Pay - Home Health",
                            "others": "Others"
                            }

        assessment_conditions = {
                                "soc": "SOC",
                                "roc": "ROC",
                                "recert": "Recert",
                                "scic": "SCIC"
                            }                   

        if file_extension == ".xlsx":
            from frappe.utils.xlsxutils import (
                read_xlsx_file_from_attached_file,
                read_xls_file_from_attached_file,
            )
            xlsx_data = read_xlsx_file_from_attached_file(file_url=import_file)
    
            modified_data = [xlsx_data[0]]
            
            payor_type_index = xlsx_data[0].index('Payor Type (as per HCHB)')
            assessment_index = xlsx_data[0].index('Assessment Type')
            
            for row in xlsx_data[1:]:
                modified_row = list(row)
                if modified_row[payor_type_index] != None:
                    modified_row[payor_type_index] = modified_row[payor_type_index].strip().replace(" ", "").lower()
                    if modified_row[payor_type_index] == "medicaid":
                        modified_row[payor_type_index] = "Medicaid"

                    if modified_row[payor_type_index] == "medicare":
                        modified_row[payor_type_index] = "Medicare"

                    for key in payor_conditions:
                        if key in modified_row[payor_type_index]:
                            modified_row[payor_type_index] = payor_conditions[key]
                
                if modified_row[assessment_index] != None:  
                    modified_row[assessment_index] = modified_row[assessment_index].strip().replace(" ", "").lower()
                    modified_row[assessment_index] = assessment_conditions[modified_row[assessment_index]]
                    

                modified_data.append(modified_row)
                
            import openpyxl
            wb = openpyxl.load_workbook(filename=file_path)
            workbook = openpyxl.load_workbook(file_path)
            sheet = wb.active
            sheet.delete_rows(1, sheet.max_row)
            sheet.delete_cols(1, sheet.max_column)
            for row_data in modified_data:
                sheet.append(row_data)
            
            wb.save(file_path)


        elif file_extension == ".csv":
            
            with open(file_path, 'r', newline='') as csvfile:
                csvreader = csv.reader(csvfile)
                data = list(csvreader)
                rows = data[1:]
                payor_type_index = data[0].index('Payor Type (as per HCHB)')
                assessment_index = data[0].index('Assessment Type')
                for row in rows:
                    if row[payor_type_index] != None:
                        row[payor_type_index] = row[payor_type_index].strip().replace(" ", "").lower()
                        if row[payor_type_index] == "medicaid":
                            row[payor_type_index] = "Medicaid"

                        if row[payor_type_index] == "medicare":
                            row[payor_type_index] = "Medicare"

                        for key in payor_conditions:
                            if key in row[payor_type_index]:
                                row[payor_type_index] = payor_conditions[key]

                    if row[assessment_index] != None:
                        row[assessment_index] = row[assessment_index].strip().replace(" ", "").lower()
                        row[assessment_index] = assessment_conditions[row[assessment_index]]
                       


            with open(file_path, 'w', newline='') as csvfile:
                csvwriter = csv.writer(csvfile)
                csvwriter.writerows(data)

                csv_content = '\n'.join([','.join(row) for row in data])
                csv_bytes = csv_content.encode('utf-8')

        else:
            return None

@frappe.whitelist()
def check_duplicate_records(data_import, import_file=None):
    if import_file:
        from frappe.utils.file_manager import save_file

        file_doc = frappe.get_doc("File", {"file_url": import_file})
        file_path = file_doc.get_full_path()

        file_extension = file_doc.get_extension()[1]

        if file_extension in (".csv", ".xlsx"):
            try:
                if file_extension == ".csv":
                    import csv
                    with open(file_path, 'r') as file:
                        reader = csv.DictReader(file)
                        data = list(reader)
                elif file_extension == ".xlsx":
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_data = {}
                        for index, cell_value in enumerate(row, start=1):
                            cell_letter = openpyxl.utils.get_column_letter(index)
                            header_value = ws.cell(row=1, column=index).value
                            row_data[header_value] = cell_value
                        data.append(row_data)

                if data:
                    duplicate_records = []

                    for row in data:
                        mr_number = row.get('MR Number') 
                        payor_type = row.get('Payor Type (as per HCHB)')
                        assessment_type = row.get('Assessment Type')
                        arrived_date = row.get('Arrived Date')

                        if arrived_date:
                            if isinstance(arrived_date, str):
                                try:
                                    given_date_object = datetime.datetime.strptime(arrived_date, "%d-%m-%Y %H:%M")
                                    given_date_object = str(given_date_object)
                                except ValueError:
                                    # Handle the case where the date format is invalid
                                    given_date_object = None
                            elif isinstance(arrived_date, datetime.datetime):
                                given_date_object = str(arrived_date)
                            else:
                                given_date_object = None
                        else:
                            given_date_object = None
                        if mr_number and payor_type and assessment_type and given_date_object:

                            # given_date_object = str(datetime.datetime.strptime(arrived_date, "%d-%m-%Y %H:%M"))
                            if frappe.db.exists("Bulk Upload Activities", {"mr_number": mr_number, "payor_type": payor_type, "assessment_type": assessment_type, "arrived_date": given_date_object}):
                                duplicate_records.append((mr_number, payor_type, assessment_type, arrived_date))
                    return duplicate_records
                else:
                    return "Data is empty"
            except FileNotFoundError:
                return "File not found"
    
    return []

